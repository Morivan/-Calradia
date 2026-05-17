from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.core.management.base import BaseCommand

from workshop.models import Client, Colleague, Material, Order, Product

HEADER_FILL = PatternFill("solid", fgColor="2C2C1E")
HEADER_FONT = Font(bold=True, color="D1B68C")


class Command(BaseCommand):
    help = "Export workshop data to XLSX files in exports/yandex-tables/"

    def handle(self, *args, **options):
        out_dir = Path("exports/yandex-tables")
        out_dir.mkdir(parents=True, exist_ok=True)

        self._export_clients(out_dir)
        self._export_orders(out_dir)
        self._export_materials(out_dir)
        self._export_colleagues(out_dir)
        self._export_products(out_dir)

        self.stdout.write(self.style.SUCCESS(f"Экспорт завершён → {out_dir.resolve()}"))

    def _write(self, path: Path, headers: list, rows) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = path.stem.split("_", 1)[-1]

        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append(list(row))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(path)
        self.stdout.write(f"  ✓ {path.name} ({path.stat().st_size // 1024 + 1} KB)")

    def _export_clients(self, out_dir: Path) -> None:
        headers = ["ID", "Имя", "VK (ссылка)", "Статус", "Заметки", "Дата добавления"]
        rows = [
            [c.id, c.name, c.vk_url, c.status, c.notes, c.created_at.strftime("%d.%m.%Y")]
            for c in Client.objects.order_by("-created_at")
        ]
        self._write(out_dir / "1_Клиенты.xlsx", headers, rows)

    def _export_orders(self, out_dir: Path) -> None:
        headers = ["ID", "ID клиента", "Имя клиента", "Изделие", "Конфигурация",
                   "Статус", "Срок сдачи", "Сумма", "Аванс", "Заметки", "Дата создания"]
        rows = [
            [
                o.id,
                o.client_id or "",
                o.client_name,
                o.product.name if o.product else o.product_name,
                o.configuration,
                o.status,
                o.deadline.strftime("%d.%m.%Y") if o.deadline else "",
                o.total,
                o.advance,
                o.notes,
                o.created_at.strftime("%d.%m.%Y"),
            ]
            for o in Order.objects.select_related("client", "product").order_by("-created_at")
        ]
        self._write(out_dir / "2_Заказы.xlsx", headers, rows)

    def _export_materials(self, out_dir: Path) -> None:
        headers = ["ID", "Название", "Тип", "Направление", "Единица", "Цена",
                   "Остаток", "Мин. остаток", "Поставщик", "Заметки"]
        rows = [
            [m.id, m.name, m.type, m.direction, m.unit,
             float(m.price) if m.price else "", float(m.stock), float(m.min_stock), m.supplier, m.notes]
            for m in Material.objects.order_by("direction", "name")
        ]
        self._write(out_dir / "3_Материалы.xlsx", headers, rows)

    def _export_colleagues(self, out_dir: Path) -> None:
        headers = ["ID", "Имя", "Направление", "Специализация", "Контакт"]
        rows = [
            [c.id, c.name, c.direction, c.specialization, c.contact]
            for c in Colleague.objects.order_by("direction", "name")
        ]
        self._write(out_dir / "6_Коллеги.xlsx", headers, rows)

    def _export_products(self, out_dir: Path) -> None:
        headers = ["ID", "Название", "Статус", "Категория", "Эпоха", "Материал",
                   "Цена от", "Срок изготовления", "Вес", "Обновлено"]
        rows = [
            [p.id, p.name, p.status, p.category, p.era, p.material,
             p.price_from, p.lead_time, p.weight, p.updated_at.strftime("%d.%m.%Y")]
            for p in Product.objects.order_by("name")
        ]
        self._write(out_dir / "7_Каталог.xlsx", headers, rows)
